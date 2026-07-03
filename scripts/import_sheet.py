#!/usr/bin/env python3
"""기존 구글시트(신규 딜러 발굴) xlsx → Supabase leads 이관 스크립트 (1회성).

사용법:
  SUPABASE_URL=https://xxx.supabase.co SUPABASE_KEY=<service_role 또는 로그인 토큰> \
    python3 import_sheet.py <시트.xlsx> [--owner Eric] [--dry-run]

단계 매핑: 승인 여부 O → 승인 / 확인 여부 O → 응답 / 컨택 여부 O → 컨택
          제외 O → 제외 / 그 외 → 발굴
"""
import argparse
import json
import os
import sys
import urllib.request

import openpyxl

HEADER_MAP = {
    "셀러명(링크)": "name",
    "발굴 채널": "channel",
    "팔로워\n/관심고객": "followers",
    "샵 상세": "shop_detail",
    "주요 상품": "main_products",
    "비고": "notes",
    "컨택포인트": "contact_point",
    "컨택일": "contact_date",
    "닉네임": "nickname",
    "승인일": "approved_at",
}


def parse_date(v):
    """'26.07.02' → '2026-07-02', datetime → ISO date."""
    if v is None or v == "":
        return None
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    parts = s.replace("-", ".").replace("/", ".").split(".")
    if len(parts) == 3:
        y, m, d = parts
        if len(y) == 2:
            y = "20" + y
        try:
            return f"{int(y):04d}-{int(m):02d}-{int(d):02d}"
        except ValueError:
            return None
    return None


def checked(v):
    return str(v).strip().upper() == "O"


def rows_from_sheet(ws, channel_type, owner, category):
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    out = []
    for row_cells in ws.iter_rows(min_row=2):
        row = [c.value for c in row_cells]
        no = str(row[idx.get("NO.", 0)] or "").strip()
        name = row[idx.get("셀러명(링크)", 1)]
        if not name or no == "예시":
            continue
        lead = {"type": "dealer", "channel_type": channel_type, "owner": owner, "category": category, "extra": {}}
        for h, key in HEADER_MAP.items():
            if h not in idx:
                continue
            v = row[idx[h]]
            if key in ("contact_date", "approved_at"):
                v = parse_date(v)
            elif key == "followers":
                try:
                    v = int(float(v)) if v not in (None, "") else None
                except ValueError:
                    v = None
            elif v is not None:
                v = str(v).strip() or None
            lead[key] = v

        # 셀러명 셀 하이퍼링크 → link
        name_cell = row_cells[idx.get("셀러명(링크)", 1)]
        if name_cell.hyperlink:
            lead["link"] = name_cell.hyperlink.target

        # 단계 산출 (우선순위: 제외 > 승인 > 확인 > 컨택 > 발굴)
        get = lambda h: row[idx[h]] if h in idx else None
        if checked(get("제외")):
            lead["stage"] = "제외"
        elif checked(get("승인 여부")):
            lead["stage"] = "승인"
        elif checked(get("확인 여부")):
            lead["stage"] = "응답"
        elif checked(get("컨택 여부")):
            lead["stage"] = "컨택"
        else:
            lead["stage"] = "발굴"

        # 진행 상황 텍스트가 있으면 비고에 병합
        prog = get("진행 상황")
        if prog and str(prog).strip():
            lead["notes"] = ((lead.get("notes") or "") + f"\n[진행 상황] {prog}").strip()

        out.append(lead)
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("--owner", default="Eric")
    ap.add_argument("--category", default="피규어")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_KEY")
    if not args.dry_run and (not url or not key):
        sys.exit("SUPABASE_URL / SUPABASE_KEY 환경변수를 설정해주세요.")

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    leads = []
    for sheet_name, channel_type in (("온라인", "온라인"), ("오프라인", "오프라인")):
        if sheet_name in wb.sheetnames:
            leads += rows_from_sheet(wb[sheet_name], channel_type, args.owner, args.category)

    print(f"파싱 완료: {len(leads)}건")
    from collections import Counter
    print("단계 분포:", dict(Counter(l["stage"] for l in leads)))

    if args.dry_run:
        print(json.dumps(leads[:3], ensure_ascii=False, indent=2))
        return

    req = urllib.request.Request(
        f"{url}/rest/v1/leads",
        data=json.dumps(leads).encode(),
        headers={
            "apikey": key,
            "Authorization": f"Bearer {os.environ.get('SUPABASE_BEARER', key)}",
            "Content-Type": "application/json",
            "Prefer": "return=minimal",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req) as r:
            print("업로드 완료:", r.status)
    except urllib.error.HTTPError as e:
        sys.exit(f"업로드 실패 {e.code}: {e.read().decode()[:500]}")


if __name__ == "__main__":
    main()
