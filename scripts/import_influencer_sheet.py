#!/usr/bin/env python3
"""인플루언서 협업 구글시트(TCG 릴스 리스트업) → Supabase leads 이관 스크립트 (1회성).

사용법:
  SUPABASE_URL=... SUPABASE_KEY=<publishable> SUPABASE_BEARER=<로그인 토큰> \
    python3 import_influencer_sheet.py <시트.xlsx> [--owner Ian] [--collab "TCG 릴스 협업"] [--dry-run]

시트 구조: 헤더 2행, 데이터 3행부터.
컬럼: NO., 전달일, 상태, 채널, 채널명, 링크, 팔로워수, 이메일, 포인트,
      연락, 응답, 계획 단가, 실제 단가, 협업 관련 특이사항, 응낙 여부, 계약 체결 여부

단계 매핑(우선순위): 상태=제외 → 제외 / 계약 체결 → 진행 / 응낙 → 계약
                  / 응답 있음 → 응답 / 연락 있음 → 컨택 / 그 외 → 발굴
"""
import argparse
import json
import os
import re
import sys
import urllib.request

import openpyxl

YEAR = 2026  # 시트에 연도가 없어 컨택 연도를 고정


def parse_followers(v):
    """'1.63만명' → 16300, '5.53천명' → 5530, '154명' → 154."""
    if v in (None, ""):
        return None
    s = str(v).replace(",", "").replace("명", "").strip()
    m = re.match(r"^([\d.]+)\s*(만|천)?$", s)
    if not m:
        return None
    n = float(m.group(1))
    unit = {"만": 10000, "천": 1000}.get(m.group(2), 1)
    return round(n * unit)


def parse_korean_date(v):
    """'7월 2일 저녁' → '2026-07-02'."""
    if v in (None, ""):
        return None
    m = re.search(r"(\d{1,2})월\s*(\d{1,2})일", str(v))
    if not m:
        return None
    return f"{YEAR:04d}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"


def filled(v):
    return v is not None and str(v).strip() not in ("", ".")


def rows_from_sheet(ws, owner, collab_type, args_category=None):
    headers = [str(c.value).strip() if c.value else "" for c in ws[2]]
    idx = {h: i for i, h in enumerate(headers)}
    out = []
    for row_cells in ws.iter_rows(min_row=3):
        row = [c.value for c in row_cells]
        name = row[idx["채널명"]]
        if not filled(name):
            continue
        get = lambda h: row[idx[h]] if h in idx else None

        # 비고: 포인트 + 협업 특이사항 병합
        notes_parts = [str(get(h)).strip() for h in ("포인트", "협업 관련 특이사항") if filled(get(h))]

        # 단계 산출
        if str(get("상태") or "").strip() == "제외":
            stage = "제외"
        elif filled(get("계약 체결 여부")):
            stage = "진행"
        elif filled(get("응낙 여부")):
            stage = "계약"
        elif filled(get("응답")):
            stage = "응답"
        elif filled(get("연락")):
            stage = "컨택"
        else:
            stage = "발굴"

        extra = {"collab_type": collab_type}
        for h, key in (("계획 단가", "planned_rate"), ("실제 단가", "actual_rate"),
                       ("연락", "contact_raw"), ("응답", "reply_raw"),
                       ("응낙 여부", "accept_raw"), ("계약 체결 여부", "contract_raw")):
            if filled(get(h)):
                extra[key] = str(get(h)).strip()

        out.append({
            "type": "influencer",
            "name": str(name).strip(),
            "link": str(get("링크")).strip() if filled(get("링크")) else None,
            "channel": str(get("채널")).strip() if filled(get("채널")) else None,
            "channel_type": "온라인",
            "followers": parse_followers(get("팔로워수")),
            "contact_point": str(get("이메일")).strip() if filled(get("이메일")) else None,
            "contact_date": parse_korean_date(get("연락")),
            "category": args_category,
            "stage": stage,
            "owner": owner,
            "notes": "\n".join(notes_parts) or None,
            "extra": extra,
        })
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("--owner", default="Danny")
    ap.add_argument("--collab", default="TCG 릴스 협업")
    ap.add_argument("--category", default=None)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_KEY")
    if not args.dry_run and (not url or not key):
        sys.exit("SUPABASE_URL / SUPABASE_KEY 환경변수를 설정해주세요.")

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    leads = rows_from_sheet(wb.worksheets[0], args.owner, args.collab, args.category)

    from collections import Counter
    print(f"파싱 완료: {len(leads)}건")
    print("단계 분포:", dict(Counter(l["stage"] for l in leads)))

    if args.dry_run:
        print(json.dumps(leads[:3], ensure_ascii=False, indent=2))
        return

    req = urllib.request.Request(
        f"{url}/rest/v1/leads",
        data=json.dumps(leads).encode(),
        headers={
            "apikey": key,
            "Authorization": f"Bearer {os.environ.get('SUPABASE_BEARER', key)}",
            "Content-Type": "application/json",
            "Prefer": "return=minimal",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req) as r:
            print("업로드 완료:", r.status)
    except urllib.error.HTTPError as e:
        sys.exit(f"업로드 실패 {e.code}: {e.read().decode()[:500]}")


if __name__ == "__main__":
    main()
